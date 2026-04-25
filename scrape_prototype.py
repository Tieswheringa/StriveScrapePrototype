import gc
import os
import re
import io
import time
import subprocess
import sys
from typing import Callable, Dict, List, Optional, Tuple

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Schrijfbare locatie op Streamlit Cloud
PLAYWRIGHT_BROWSERS_DIR = "/tmp/pw-browsers"
os.environ["PLAYWRIGHT_BROWSERS_PATH"] = PLAYWRIGHT_BROWSERS_DIR

# ─── Instellingen ─────────────────────────────────────────────────────────────
MEMORY_COOLDOWN_INTERVAL = 5    # Pauze na elke N opdrachten
MEMORY_COOLDOWN_SECONDEN = 90   # Duur van de geheugen-pauze
BATCH_GROOTTE = 3               # Opdrachten per batch
COOLDOWN_SECONDEN = 3           # Wachttijd tussen opdrachten


# ─── Chromium-processen forceren afsluiten ────────────────────────────────────
def kill_chromium():
    try:
        subprocess.run(["pkill", "-f", "chromium"], capture_output=True)
        time.sleep(1)
    except Exception:
        pass


@st.cache_resource(show_spinner="Playwright-browsers installeren (eenmalig)...")
def installeer_playwright():
    os.makedirs(PLAYWRIGHT_BROWSERS_DIR, exist_ok=True)
    try:
        result = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            capture_output=True,
            text=True,
            timeout=300,
            env={**os.environ, "PLAYWRIGHT_BROWSERS_PATH": PLAYWRIGHT_BROWSERS_DIR},
        )
        return {
            "returncode": result.returncode,
            "stdout": result.stdout,
            "stderr": result.stderr,
        }
    except Exception as e:
        return {"returncode": 1, "stdout": "", "stderr": str(e)}


_playwright_status = installeer_playwright()

if _playwright_status["returncode"] != 0:
    st.error("Failed to install browsers")
    st.code(_playwright_status["stderr"] or _playwright_status["stdout"])
    st.stop()

# ─── Pagina-configuratie ───────────────────────────────────────────────────────
st.set_page_config(
    page_title="Striive Matcher",
    page_icon="⚡",
    layout="wide",
)

# ─── CSS ──────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600;700&display=swap');

html, body, [class*="css"] {
    font-family: 'IBM Plex Sans', sans-serif;
}
.stApp { background-color: #0f1117; color: #e0e0e0; }
section[data-testid="stSidebar"] {
    background-color: #161b22;
    border-right: 1px solid #2d333b;
}
h1, h2, h3 { font-family: 'IBM Plex Mono', monospace !important; letter-spacing: -0.5px; }
.stButton > button {
    background-color: #238636; color: #ffffff; border: none;
    border-radius: 6px; font-family: 'IBM Plex Mono', monospace;
    font-weight: 600; letter-spacing: 0.5px; padding: 0.6rem 1.4rem;
    transition: background-color 0.2s ease;
}
.stButton > button:hover { background-color: #2ea043; }
div[data-testid="metric-container"] {
    background-color: #161b22; border: 1px solid #2d333b;
    border-radius: 8px; padding: 1rem 1.2rem;
}
.stDataFrame { border: 1px solid #2d333b; border-radius: 8px; }
.log-box {
    background-color: #0d1117; border: 1px solid #2d333b; border-radius: 8px;
    padding: 1rem; font-family: 'IBM Plex Mono', monospace; font-size: 12px;
    color: #8b949e; height: 260px; overflow-y: auto; white-space: pre-wrap;
}
.stSlider label { font-family: 'IBM Plex Mono', monospace; font-size: 13px; color: #8b949e; }
.stDownloadButton > button {
    background-color: #1f6feb; color: #ffffff; border: none;
    border-radius: 6px; font-family: 'IBM Plex Mono', monospace; font-weight: 600;
}
.stDownloadButton > button:hover { background-color: #388bfd; }
hr { border-color: #2d333b; }
</style>
""", unsafe_allow_html=True)

# ─── Hulpfuncties ─────────────────────────────────────────────────────────────

def normaliseer_url(url: str) -> str:
    return (url or "").strip().rstrip("/")


def extraheer_uurtarief(tekst: str) -> str:
    patronen = [
        r'[Uu]urtarief[:\s]*[€]?\s*(\d+[\.,]?\d*)',
        r'[Tt]arief[:\s]*[€]?\s*(\d+[\.,]?\d*)',
        r'[€]\s*(\d+[\.,]?\d*)\s*per uur',
        r'(\d+[\.,]?\d*)\s*[€]?\s*per uur',
        r'[Hh]ourly rate[:\s]*[€$]?\s*(\d+[\.,]?\d*)',
        r'[Rr]ate[:\s]*[€$]?\s*(\d+[\.,]?\d*)',
    ]
    for patroon in patronen:
        match = re.search(patroon, tekst)
        if match:
            return f"€{match.group(1)}/uur"
    return "-"


def extraheer_startdatum(tekst: str) -> str:
    patronen = [
        r'[Ss]tartdatum[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
        r'[Ss]tart[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
        r'[Uu]iterlijk[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
        r'[Vv]oor\s+(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
        r'(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
    ]
    for patroon in patronen:
        match = re.search(patroon, tekst)
        if match:
            return match.group(1)
    return "-"


def extraheer_reageer_deadline(tekst: str) -> str:
    patronen = [
        r'[Rr]eageren kan t/m\s+(\d{1,2}[-/]\d{1,2}[-/]\d{2,4})',
        r'[Rr]eageren kan t/m\s+(\d{1,2}\s+\w+\s+\d{4})',
        r'[Rr]eageren kan t/m\s+([^\n]+)',
    ]
    for patroon in patronen:
        match = re.search(patroon, tekst)
        if match:
            return match.group(1).strip()
    return "-"


def maak_excel(matches: List[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"

    headers = [
        "Opdracht #", "Kandidaat", "Score", "Uurtarief",
        "Startdatum", "Reageren t/m", "Link naar opdracht", "CV herschrijven",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial")
        cell.fill = PatternFill("solid", start_color="2E4057")
        cell.alignment = Alignment(horizontal="center")

    breedtes = [12, 25, 10, 15, 18, 18, 50, 50]
    for i, breedte in enumerate(breedtes, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = breedte

    for row_index, match in enumerate(matches, 2):
        waarden = [
            match["opdracht"], match["naam"], match["score"],
            match["uurtarief"], match["startdatum"], match["deadline"],
            match["url"],
            "https://chatgpt.com/g/g-692562722fd48191a45a59eef67f00f2-inthearena-cv-builder",
        ]
        for col, waarde in enumerate(waarden, 1):
            cell = ws.cell(row=row_index, column=col, value=waarde)
            cell.font = Font(name="Arial")
            cell.fill = PatternFill("solid", start_color="E8F5E9")

        link_cell = ws.cell(row=row_index, column=7, value=match["url"])
        link_cell.hyperlink = match["url"]
        link_cell.font = Font(name="Arial", color="0563C1", underline="single")

        cv_cell = ws.cell(row=row_index, column=8, value="Open CV Builder")
        cv_cell.hyperlink = "https://chatgpt.com/g/g-692562722fd48191a45a59eef67f00f2-inthearena-cv-builder"
        cv_cell.font = Font(name="Arial", color="0563C1", underline="single")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def run_scraper(
    credentials: Dict[str, str],
    drempel: int,
    start_bij_link: str,
    stop_bij_link: str,
    log_fn: Callable[[str], None],
    progress_fn: Callable[[int, int], None],
    result_fn: Callable[[List[dict]], None],
    batch_done_fn: Optional[Callable[[], None]] = None,
):
    from playwright.sync_api import sync_playwright

    def log(msg: str):
        log_fn(msg)

    def chunks(lst, size):
        for i in range(0, len(lst), size):
            yield lst[i:i + size]

    def sluit_browser_veilig(browser, naam: str = "browser"):
        if browser is None:
            return
        try:
            browser.close()
            log(f"  📕 {naam} gesloten.")
        except Exception as e:
            log(f"  ⚠️ Fout bij sluiten {naam}: {e}")

    def geheugen_opruimen(verwerkt: int):
        kill_chromium()
        gc.collect()
        log(
            f"\n🧹 Geheugen vrijgemaakt na {verwerkt} opdrachten. "
            f"Pauze van {MEMORY_COOLDOWN_SECONDEN}s...\n"
        )
        time.sleep(MEMORY_COOLDOWN_SECONDEN)
        gc.collect()
        kill_chromium()
        log("▶️  Hervatten na geheugen-pauze.\n")

    CHROMIUM_ARGS = [
        "--no-sandbox",
        "--disable-dev-shm-usage",
        "--disable-gpu",
        "--disable-extensions",
        "--disable-background-networking",
        "--disable-sync",
        "--disable-translate",
        "--disable-default-apps",
        "--mute-audio",
        "--no-first-run",
        "--safebrowsing-disable-auto-update",
        "--js-flags=--max-old-space-size=256",
    ]

    def login_striive(page):
        log("🔐 Inloggen op Striive...")
        page.goto("https://login.striive.com/", wait_until="domcontentloaded", timeout=60000)
        page.wait_for_timeout(3000)

        email_selectors = [
            'input[type="email"]', 'input[name="email"]', '#email',
            'input[id="email"]', 'input[placeholder*="mail" i]',
            'input[autocomplete="email"]', 'input[autocomplete="username"]',
        ]

        email_veld = None
        for selector in email_selectors:
            try:
                element = page.locator(selector).first
                element.wait_for(state="visible", timeout=5000)
                email_veld = element
                log(f"  ✔ E-mailveld gevonden: {selector}")
                break
            except Exception:
                continue

        if email_veld is None:
            try:
                page.screenshot(path="/tmp/debug_login.png", full_page=True)
            except Exception:
                pass
            raise Exception("E-mailveld niet gevonden.")

        email_veld.click()
        email_veld.fill(credentials["email"])

        wachtwoord_veld = page.locator('input[type="password"]').first
        wachtwoord_veld.wait_for(state="visible", timeout=10000)
        wachtwoord_veld.click()
        wachtwoord_veld.fill(credentials["wachtwoord"])

        login_selectors = [
            'button:has-text("Login")', 'button:has-text("Inloggen")',
            'button[type="submit"]', 'input[type="submit"]',
        ]
        for selector in login_selectors:
            try:
                knop = page.locator(selector).first
                knop.wait_for(state="visible", timeout=4000)
                knop.click()
                log(f"  Login-knop geklikt ({selector}).")
                break
            except Exception:
                continue
        else:
            wachtwoord_veld.press("Enter")
            log("  Login-knop niet gevonden, Enter gebruikt als fallback.")

        try:
            page.wait_for_selector('text=Overzicht', timeout=30000)
        except Exception:
            page.wait_for_url("**/dashboard**", timeout=30000)

        log("✅ Ingelogd!")

    def ga_naar_opdrachten(page):
        try:
            page.click('a:has-text("Opdrachten")', timeout=10000)
        except Exception:
            page.goto("https://supplier.striive.com/job-requests", wait_until="domcontentloaded", timeout=60000)

        page.wait_for_selector('[data-testid="jobRequestListItem"]', timeout=30000)
        page.wait_for_timeout(2000)
        log("✅ Opdrachtenpagina geladen.")

    def verzamel_opdracht_urls(
        page,
        start_bij_link: Optional[str] = None,
        stop_bij_link: Optional[str] = None,
    ) -> List[str]:
        start_bij_link_norm = normaliseer_url(start_bij_link)
        stop_bij_link_norm = normaliseer_url(stop_bij_link)
        gebruik_start_link = bool(start_bij_link_norm)
        gebruik_stop_link = bool(stop_bij_link_norm)

        if gebruik_start_link:
            log(f"▶️  Start-link ingesteld: {start_bij_link_norm}")
        if gebruik_stop_link:
            log(f"⛔ Stop-link ingesteld: {stop_bij_link_norm}")
        if not gebruik_start_link and not gebruik_stop_link:
            log("🔍 Alle opdracht-URLs verzamelen via scrollen...")

        # Verzamel eerst alle URLs (we moeten de volledige lijst kennen om te knippen)
        alle_urls = []
        gevonden_set = set()
        stop_link_gevonden = False
        scroll_stap = 150
        scroll_pos = 0
        geen_nieuw = 0
        max_geen_nieuw = 8

        while geen_nieuw < max_geen_nieuw and not stop_link_gevonden:
            items = page.locator('[data-testid="jobRequestListItem"]').all()
            aantal_voor = len(gevonden_set)

            for item in items:
                try:
                    href = item.get_attribute('href')
                    if not href:
                        href = item.locator('a').first.get_attribute('href')
                    if not href:
                        continue

                    volledige_href = (
                        f"https://supplier.striive.com{href}"
                        if href.startswith("/") else href
                    )
                    volledige_href_norm = normaliseer_url(volledige_href)

                    if volledige_href_norm == stop_bij_link_norm and gebruik_stop_link:
                        stop_link_gevonden = True
                        log("🛑 Stop-link bereikt.")
                        break

                    if volledige_href_norm not in gevonden_set:
                        gevonden_set.add(volledige_href_norm)
                        alle_urls.append(volledige_href_norm)

                except Exception:
                    continue

            if stop_link_gevonden:
                break

            scroll_pos += scroll_stap
            try:
                res = page.evaluate(f"""
                    () => {{
                        const s = document.querySelector('div.p-scroller');
                        if (s) {{
                            s.scrollTop = {scroll_pos};
                            return {{scrollTop: s.scrollTop, scrollHeight: s.scrollHeight, clientHeight: s.clientHeight}};
                        }}
                        window.scrollTo(0, {scroll_pos});
                        return {{scrollTop: window.scrollY, scrollHeight: document.body.scrollHeight, clientHeight: window.innerHeight}};
                    }}
                """)
            except Exception:
                res = {"scrollTop": 0, "scrollHeight": 0, "clientHeight": 0}

            page.wait_for_timeout(1000)

            if len(gevonden_set) > aantal_voor:
                geen_nieuw = 0
            else:
                geen_nieuw += 1

            max_scroll = res["scrollHeight"] - res["clientHeight"]
            if max_scroll > 0 and res["scrollTop"] >= max_scroll - 10:
                log(f"📋 Einde van lijst bereikt. Totaal: {len(alle_urls)} gevonden.")
                break

        if gebruik_stop_link and not stop_link_gevonden:
            log("⚠️ Stop-link niet gevonden in de lijst.")

        # ── Snij de lijst bij op basis van start- en/of stop-link ────────────
        start_index = 0
        if gebruik_start_link:
            try:
                start_index = alle_urls.index(start_bij_link_norm)
                log(f"▶️  Start-link gevonden op positie {start_index + 1}.")
            except ValueError:
                log("⚠️ Start-link niet gevonden. Verwerking begint vanaf het begin.")
                start_index = 0

        geselecteerd = alle_urls[start_index:]

        # Als stop-link gevonden is, zijn de URLs tot aan de stop al geselecteerd
        # (stop-link zelf is niet toegevoegd aan alle_urls, dus de lijst eindigt correct)
        log(f"📋 Totaal {len(geselecteerd)} opdrachten geselecteerd.")
        return geselecteerd

    def veilige_inner_texts(locator) -> List[str]:
        try:
            return locator.all_inner_texts()
        except Exception:
            return []

    def login_streamlit_eenmalig(page, frame):
        page.goto(
            "https://inthearenabv-cv-tool.streamlit.app/",
            wait_until="domcontentloaded",
            timeout=60000,
        )
        page.wait_for_timeout(20000)

        try:
            wachtwoord_veld = frame.locator('input[type="password"]').first
            if wachtwoord_veld.is_visible(timeout=5000):
                wachtwoord_veld.fill(credentials["streamlit_pw"])
                frame.locator('button:has-text("Log in")').first.click()
                page.wait_for_timeout(12000)
                log("  🔑 Streamlit ingelogd.")
        except Exception:
            pass

        try:
            tab = frame.locator('button:has-text("Test geschiktheid opdracht")').first
            if tab.is_visible(timeout=5000):
                tab.click()
                page.wait_for_timeout(2500)
                log("  🖱️ Tab geklikt.")
        except Exception:
            pass

    def analyseer_met_bestaande_pagina(page, frame, tekst: str) -> List[Tuple[str, int]]:
        textarea = frame.locator("textarea").first
        textarea.wait_for(state="visible", timeout=30000)

        oude_resultaat_tekst = "\n".join(
            veilige_inner_texts(frame.locator('[data-testid="stExpander"]'))
        ).strip()

        textarea.click(timeout=10000)
        try:
            page.keyboard.press("Control+A")
        except Exception:
            pass
        try:
            page.keyboard.press("Backspace")
        except Exception:
            pass
        page.wait_for_timeout(300)
        textarea.fill("")
        page.wait_for_timeout(300)
        textarea.fill(tekst)
        page.wait_for_timeout(800)

        ingevulde_tekst = ""
        try:
            ingevulde_tekst = textarea.input_value(timeout=5000)
        except Exception:
            pass

        if not ingevulde_tekst or len(ingevulde_tekst.strip()) < 50:
            raise Exception("Textarea lijkt niet correct gevuld.")

        log("  ✍️ Tekst ingevuld.")

        analyse_knop = frame.locator('button:has-text("Analyseer geschiktheid")').first
        analyse_knop.wait_for(state="visible", timeout=10000)
        analyse_knop.click()
        log("  ⏳ Analyse gestart...")

        # Wacht op spinner
        spinner_verschenen = False
        start = time.time()
        while time.time() - start < 15:
            try:
                if frame.locator('[data-testid="stSpinner"]').count() > 0:
                    spinner_verschenen = True
                    log("  🔄 Spinner gedetecteerd.")
                    break
            except Exception:
                pass
            page.wait_for_timeout(500)

        if spinner_verschenen:
            start = time.time()
            while time.time() - start < 120:
                try:
                    if frame.locator('[data-testid="stSpinner"]').count() == 0:
                        log("  ✅ Spinner verdwenen.")
                        break
                except Exception:
                    break
                page.wait_for_timeout(1000)

        # Wacht op stabiele nieuwe resultaten
        laatste_tekst = ""
        stabiele_teller = 0
        nieuwe_resultaat_tekst = ""

        start = time.time()
        while time.time() - start < 120:
            try:
                expanders = frame.locator('[data-testid="stExpander"]')
                if expanders.count() > 0:
                    huidige_tekst = "\n".join(veilige_inner_texts(expanders)).strip()
                    if huidige_tekst and huidige_tekst != oude_resultaat_tekst:
                        if huidige_tekst == laatste_tekst:
                            stabiele_teller += 1
                        else:
                            stabiele_teller = 0
                        laatste_tekst = huidige_tekst
                        if stabiele_teller >= 2:
                            nieuwe_resultaat_tekst = huidige_tekst
                            break
            except Exception:
                pass
            page.wait_for_timeout(1200)

        if not nieuwe_resultaat_tekst:
            raise Exception("Resultaten niet stabiel of niet vernieuwd binnen timeout.")

        page.wait_for_timeout(1000)

        resultaten: List[Tuple[str, int]] = []
        aantal_expanders = frame.locator('[data-testid="stExpander"]').count()

        for i in range(aantal_expanders):
            try:
                blok = frame.locator('[data-testid="stExpander"]').nth(i)
                try:
                    blok.locator('summary, [data-testid="stExpanderToggleIcon"], button').first.click(timeout=3000)
                    page.wait_for_timeout(300)
                except Exception:
                    pass

                blok_tekst = blok.inner_text(timeout=15000)
                score_match = re.search(r'(\d+)/100', blok_tekst)
                if not score_match:
                    continue

                score = int(score_match.group(1))
                naam_match = re.search(r'[🟢🟡🔴]\s*(.*?)\s*—\s*\d+/100', blok_tekst)
                naam = naam_match.group(1).strip() if naam_match else "onbekend"
                resultaten.append((naam, score))
                log(f"  👤 {naam} → {score}/100")

            except Exception as e:
                log(f"  ⚠️ Kandidaatblok fout: {e}")

        if not resultaten:
            raise Exception("Geen geldige kandidaten gevonden na analyse.")

        return resultaten

    def analyseer_met_timeout(page, frame, tekst: str) -> List[Tuple[str, int]]:
        """
        Roept analyseer_met_bestaande_pagina direct aan (greenlet-safe).
        Ingebouwde Playwright-timeouts (30s, 120s) voorkomen eindeloos wachten.
        """
        try:
            return analyseer_met_bestaande_pagina(page, frame, tekst)
        except Exception as e:
            log(f"  ⚠️ Analyse-fout: {e}")
            return []

    # ─── Hoofd scraper loop ───────────────────────────────────────────────────
    alle_matches = []
    mislukte_urls = []

    # ── Stap 1: URLs verzamelen via tijdelijke init-browser ───────────────────
    alle_urls = []
    with sync_playwright() as p:
        init_browser = None
        try:
            init_browser = p.chromium.launch(headless=True, args=CHROMIUM_ARGS)
            init_page = init_browser.new_page(viewport={"width": 1280, "height": 800})
            init_page.on("crash", lambda: log("💥 Init-page crash"))
            login_striive(init_page)
            ga_naar_opdrachten(init_page)
            alle_urls = verzamel_opdracht_urls(
                init_page,
                start_bij_link=start_bij_link,
                stop_bij_link=stop_bij_link,
            )
        finally:
            sluit_browser_veilig(init_browser, "init-browser")
            gc.collect()
            kill_chromium()

    if not alle_urls:
        log("⚠️ Geen nieuwe opdrachten gevonden.")
        return []

    totaal = len(alle_urls)
    verwerkt = 0

    # ── Stap 2: Elke batch in een eigen sync_playwright context ───────────────
    for batch_nummer, batch_urls in enumerate(chunks(alle_urls, BATCH_GROOTTE), start=1):
        log(f"\n📦 Start batch {batch_nummer} ({len(batch_urls)} opdrachten)")

        with sync_playwright() as p:
            batch_browser = None
            streamlit_browser = None
            streamlit_context = None
            striive_page = None
            streamlit_page = None

            try:
                # Striive-browser voor deze batch
                batch_browser = p.chromium.launch(headless=True, args=CHROMIUM_ARGS)
                batch_browser.on("disconnected", lambda: log(f"🔌 Batch-browser {batch_nummer} disconnected"))
                striive_page = batch_browser.new_page(viewport={"width": 1280, "height": 800})
                striive_page.on("crash", lambda: log(f"💥 Striive-page crash batch {batch_nummer}"))
                login_striive(striive_page)

                # Streamlit-browser eenmalig openen per batch
                streamlit_browser = p.chromium.launch(headless=True, args=CHROMIUM_ARGS)
                streamlit_context = streamlit_browser.new_context(viewport={"width": 1280, "height": 900})
                streamlit_page = streamlit_context.new_page()
                streamlit_frame = streamlit_page.frame_locator("iframe").first

                try:
                    login_streamlit_eenmalig(streamlit_page, streamlit_frame)
                    log(f"  🌐 Streamlit-sessie klaar voor batch {batch_nummer}.")
                except Exception as e:
                    log(f"  ⚠️ Streamlit login mislukt voor batch {batch_nummer}: {e}")

                # Opdrachten in deze batch verwerken
                for url in batch_urls:
                    verwerkt += 1
                    progress_fn(verwerkt, totaal)
                    log(f"\n[{verwerkt}/{totaal}] {url}")

                    try:
                        striive_page.goto(url, wait_until="domcontentloaded", timeout=60000)
                        striive_page.wait_for_timeout(2000)
                        tekst = striive_page.locator("app-job-request-details").inner_text(timeout=15000)

                        uurtarief = extraheer_uurtarief(tekst)
                        startdatum = extraheer_startdatum(tekst)
                        deadline = extraheer_reageer_deadline(tekst)
                        log(f"  💶 {uurtarief} | 📅 {startdatum} | ⏰ {deadline}")

                        time.sleep(COOLDOWN_SECONDEN)

                        kandidaten = analyseer_met_timeout(streamlit_page, streamlit_frame, tekst)

                        toegevoegde_kandidaten = set()
                        for naam, score in kandidaten:
                            unieke_sleutel = (url, naam, score)
                            if score > drempel and unieke_sleutel not in toegevoegde_kandidaten:
                                toegevoegde_kandidaten.add(unieke_sleutel)
                                alle_matches.append({
                                    "opdracht": f"Opdracht {verwerkt}",
                                    "naam": naam,
                                    "score": score,
                                    "uurtarief": uurtarief,
                                    "startdatum": startdatum,
                                    "deadline": deadline,
                                    "url": url,
                                })
                                log(f"  ✅ Match! {naam} ({score}/100) boven drempel {drempel}.")
                                result_fn(alle_matches)

                    except Exception as e:
                        log(f"  ⚠️ Opdracht overgeslagen: {e}")
                        mislukte_urls.append(url)
                        continue

            except Exception as e:
                log(f"❌ Batch {batch_nummer} gecrasht: {e}")
                for url in batch_urls:
                    if url not in mislukte_urls:
                        mislukte_urls.append(url)

            finally:
                sluit_browser_veilig(streamlit_browser, f"streamlit-browser-{batch_nummer}")
                sluit_browser_veilig(batch_browser, f"batch-browser-{batch_nummer}")
                gc.collect()
                kill_chromium()
                log(f"📦 Batch {batch_nummer} afgesloten + Chromium opgeruimd.")

                if batch_done_fn:
                    batch_done_fn()

        # Geheugen-pauze buiten de with-context (Playwright-server is al gestopt)
        if verwerkt % MEMORY_COOLDOWN_INTERVAL == 0 and verwerkt < totaal:
            geheugen_opruimen(verwerkt)

    if mislukte_urls:
        log(f"\n⚠️ {len(mislukte_urls)} opdrachten mislukt of overgeslagen.")

    log(f"\n🏁 Klaar! {len(alle_matches)} matches gevonden.")
    return alle_matches


# ─── Session State initialisatie ──────────────────────────────────────────────
defaults = {
    "matches": [],
    "logs": [],
    "bezig": False,
    "klaar": False,
    "voortgang": (0, 0),
}
for key, default in defaults.items():
    if key not in st.session_state:
        st.session_state[key] = default

# ─── UI ───────────────────────────────────────────────────────────────────────
col_logo, col_titel = st.columns([1, 8])

with col_logo:
    st.markdown("<div style='font-size:48px;padding-top:8px'>⚡</div>", unsafe_allow_html=True)

with col_titel:
    st.markdown("""
        <h1 style='margin:0;padding-top:12px;color:#e6edf3'>Striive Matcher</h1>
        <p style='color:#8b949e;margin:0;font-size:14px'>
            Automatisch opdrachten ophalen, analyseren en de beste kandidaten vinden.
        </p>
    """, unsafe_allow_html=True)

st.markdown("---")

# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("### 🔐 Inloggegevens")
    email = st.text_input("Striive e-mailadres", value="info@inthearena.nl")
    ww = st.text_input("Striive wachtwoord", type="password", value="InTheArena2@22")

    st.markdown("---")
    st.markdown("### 🤖 CV-tool")
    streamlit_pw = st.text_input("Streamlit wachtwoord", type="password", value="InTheArenaBV")

    st.markdown("---")
    st.markdown("### 🧷 Bereik instellen")
    start_bij_link = st.text_input(
        "Beginopdracht (inclusief)",
        value="",
        placeholder="https://supplier.striive.com/inbox/all/...",
        help=(
            "Plak hier de link van de opdracht waarmee je wilt beginnen. "
            "Laat leeg om vanaf de eerste opdracht te starten."
        ),
    )
    stop_bij_link = st.text_input(
        "Eindopdracht (exclusief)",
        value="",
        placeholder="https://supplier.striive.com/inbox/all/...",
        help=(
            "Plak hier de link van de opdracht waarbij je wilt stoppen (deze wordt niet meer meegenomen). "
            "Laat leeg om tot de laatste opdracht door te gaan."
        ),
    )
    if start_bij_link or stop_bij_link:
        st.caption(
            f"📌 Bereik: "
            f"{'vanaf opgegeven start' if start_bij_link else 'begin'}"
            f" → "
            f"{'t/m opgegeven stop' if stop_bij_link else 'einde'}"
        )

    st.markdown("---")
    st.markdown("### ⚙️ Instellingen")
    drempel = st.slider(
        "Minimale score",
        min_value=50, max_value=95, value=80, step=5,
        help="Alleen kandidaten boven deze score worden opgenomen.",
    )

    st.markdown("---")
    if isinstance(_playwright_status, dict) and _playwright_status.get("returncode") == 0:
        st.success("✅ Playwright Chromium klaar.", icon="✅")
    else:
        st.error("❌ Playwright-installatie mislukt.")
        if isinstance(_playwright_status, dict):
            st.code(_playwright_status.get("stderr") or _playwright_status.get("stdout") or "Onbekende fout")
        else:
            st.code(str(_playwright_status))

    st.caption("v2.1 · In The Arena BV")

# ─── Hoofd kolommen ───────────────────────────────────────────────────────────
col_links, col_rechts = st.columns([3, 2], gap="large")

with col_links:
    st.markdown("### 📊 Resultaten")
    metrics_placeholder = st.empty()
    resultaat_placeholder = st.empty()

with col_rechts:
    st.markdown("### 📋 Live log")
    log_placeholder = st.empty()

st.markdown("---")
col_btn, col_status = st.columns([2, 5])

with col_btn:
    start_knop = st.button(
        "🚀  Start analyse",
        disabled=st.session_state.bezig,
        use_container_width=True,
    )

with col_status:
    progress_placeholder = st.empty()


def render_resultaten():
    import pandas as pd

    with metrics_placeholder.container():
        metric_1, metric_2, metric_3 = st.columns(3)
        verwerkt, totaal = st.session_state.voortgang
        metric_1.metric("Verwerkt", f"{verwerkt}/{totaal}" if totaal else "0/0")
        metric_2.metric("Matches", len(st.session_state.matches))
        metric_3.metric("Drempelwaarde", f"{drempel}/100")

    resultaat_placeholder.empty()

    with resultaat_placeholder.container():
        if st.session_state.matches:
            df = pd.DataFrame(st.session_state.matches)
            df_weergave = df[["opdracht", "naam", "score", "uurtarief", "startdatum", "deadline"]].copy()
            df_weergave.columns = ["Opdracht", "Kandidaat", "Score", "Uurtarief", "Startdatum", "Reageren t/m"]
            df_weergave["CV Herschrijven"] = "https://chatgpt.com/g/g-692562722fd48191a45a59eef67f00f2-inthearena-cv-builder"

            st.dataframe(
                df_weergave,
                column_config={
                    "CV Herschrijven": st.column_config.LinkColumn(
                        "CV Herschrijven",
                        display_text="✏️ Open CV Builder",
                    )
                },
                use_container_width=True,
                hide_index=True,
            )

            if not st.session_state.bezig:
                excel_bytes = maak_excel(st.session_state.matches)
                st.download_button(
                    label="⬇️  Download als Excel",
                    data=excel_bytes,
                    file_name="striive_matches.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_final",
                )

        elif st.session_state.klaar:
            st.info("Geen matches gevonden boven de ingestelde drempel.")
        else:
            st.markdown(
                "<p style='color:#8b949e;font-size:14px'>Nog geen resultaten. Start de analyse via de knop hieronder.</p>",
                unsafe_allow_html=True,
            )


def render_log():
    log_tekst = "\n".join(st.session_state.logs[-80:])
    log_placeholder.markdown(
        f"<div class='log-box'>{log_tekst}</div>",
        unsafe_allow_html=True,
    )


def render_progress():
    verwerkt, totaal = st.session_state.voortgang
    if st.session_state.bezig:
        if totaal:
            progress_placeholder.progress(
                verwerkt / totaal,
                text=f"Verwerken {verwerkt} van {totaal} opdrachten...",
            )
        else:
            progress_placeholder.info("Bezig met opstarten...")
    else:
        progress_placeholder.empty()


render_resultaten()
render_log()
render_progress()

# ─── Scraper starten ──────────────────────────────────────────────────────────
if start_knop:
    if not email or not ww:
        st.error("Vul eerst je Striive-inloggegevens in.")
    else:
        st.session_state.bezig = True
        st.session_state.klaar = False
        st.session_state.matches = []
        st.session_state.logs = []
        st.session_state.voortgang = (0, 0)

        render_resultaten()
        render_log()
        render_progress()

        def log_fn(msg):
            st.session_state.logs.append(msg)
            render_log()

        def progress_fn(huidig, totaal):
            st.session_state.voortgang = (huidig, totaal)
            render_progress()
            render_resultaten()

        def result_fn(matches):
            st.session_state.matches = matches
            render_resultaten()

        def batch_done_fn():
            render_log()
            render_progress()
            render_resultaten()

        credentials = {
            "email": email,
            "wachtwoord": ww,
            "streamlit_pw": streamlit_pw,
        }

        try:
            matches = run_scraper(
                credentials=credentials,
                drempel=drempel,
                start_bij_link=start_bij_link,
                stop_bij_link=stop_bij_link,
                log_fn=log_fn,
                progress_fn=progress_fn,
                result_fn=result_fn,
                batch_done_fn=batch_done_fn,
            )
            st.session_state.matches = matches
            st.session_state.klaar = True
        except Exception as e:
            st.session_state.logs.append(f"\n❌ Fout: {e}")
            render_log()
        finally:
            st.session_state.bezig = False
            render_progress()
            render_resultaten()

        st.rerun()
